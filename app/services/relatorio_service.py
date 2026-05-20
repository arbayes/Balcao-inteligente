"""
Serviço de Relatórios - Análises e Estatísticas do Sistema
"""

from datetime import datetime
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database.clientes_repository import listar_clientes
from app.database.estoque_repository import listar_produtos
from app.database.categorias_repository import obter_margem_categoria_por_nome


def _gerar_resumo_caixa():
    try:
        from app.services.caixa_service import obter_resumo_caixa, historico_caixas

        caixa_aberto = obter_resumo_caixa()
        caixas = historico_caixas(10)
        fechados = [c for c in caixas if c.get("status") == "FECHADO"]
        diferenca_total = sum(float(c.get("diferenca") or 0) for c in fechados)

        return {
            "aberto": bool(caixa_aberto),
            "atual": caixa_aberto,
            "historico": caixas,
            "ultimo_fechado": fechados[0] if fechados else None,
            "diferenca_total_fechados": diferenca_total,
            "quantidade_fechados": len(fechados),
        }
    except Exception as e:
        return {
            "aberto": False,
            "atual": None,
            "historico": [],
            "ultimo_fechado": None,
            "diferenca_total_fechados": 0,
            "quantidade_fechados": 0,
            "erro": str(e),
        }


def gerar_relatorio_geral() -> dict:
    """
    Gera relatório geral do sistema.
    
    Returns:
        dict: Dados agregados do sistema
    """
    clientes = listar_clientes()
    produtos = listar_produtos()
    
    # Contadores
    total_clientes = len(clientes)
    total_produtos = len(produtos)
    
    # Cálculos de estoque
    total_itens_estoque = sum(p.quantidade for p in produtos)
    valor_total_custo = sum(p.preco_compra * p.quantidade for p in produtos)
    valor_total_venda = sum(p.preco_venda * p.quantidade for p in produtos)
    lucro_potencial = valor_total_venda - valor_total_custo
    
    # Produtos com baixo estoque
    baixo_estoque = [p for p in produtos if p.quantidade <= 5]
    
    # Produtos com alta margem
    alta_margem = [p for p in produtos if p.margem_lucro >= 50]
    
    # Produtos com baixa margem
    baixa_margem = [p for p in produtos if p.margem_lucro < 20]

    # Produtos abaixo da margem alvo (produto > categoria > none)
    produtos_abaixo_margem_alvo = []
    for p in produtos:
        margem_alvo = p.margem_alvo
        if margem_alvo is None and p.categoria:
            margem_alvo = obter_margem_categoria_por_nome(p.categoria)

        if margem_alvo is None:
            continue

        margem_atual = p.margem_lucro
        margem_alvo_percent = margem_alvo * 100
        if margem_atual + 1e-6 < margem_alvo_percent:
            produtos_abaixo_margem_alvo.append({
                "nome": p.nome,
                "sku": p.sku,
                "categoria": p.categoria,
                "margem_atual": margem_atual,
                "margem_alvo": margem_alvo_percent,
                "preco_compra": p.preco_compra,
                "preco_venda": p.preco_venda
            })

    produtos_abaixo_margem_alvo.sort(
        key=lambda item: (item["margem_alvo"] - item["margem_atual"]),
        reverse=True
    )
    
    return {
        "total_clientes": total_clientes,
        "total_produtos": total_produtos,
        "total_itens_estoque": total_itens_estoque,
        "valor_total_custo": valor_total_custo,
        "valor_total_venda": valor_total_venda,
        "lucro_potencial": lucro_potencial,
        "margem_media": sum(p.margem_lucro for p in produtos) / len(produtos) if produtos else 0,
        "baixo_estoque": baixo_estoque,
        "alta_margem": alta_margem,
        "baixa_margem": baixa_margem,
        "produtos_abaixo_margem_alvo": produtos_abaixo_margem_alvo,
        "produtos": produtos,
        "caixa": _gerar_resumo_caixa()
    }


def gerar_recomendacoes(relatorio: dict) -> list:
    """
    Gera recomendações baseado no relatório.
    
    Args:
        relatorio (dict): Relatório geral
    
    Returns:
        list: Lista de recomendações
    """
    recomendacoes = []
    
    # Verificar produtos com baixo estoque
    if relatorio["baixo_estoque"]:
        recomendacoes.append({
            "tipo": "alerta",
            "titulo": "⚠️ PRODUTOS COM BAIXO ESTOQUE",
            "mensagem": f"Você tem {len(relatorio['baixo_estoque'])} produto(s) com quantidade ≤ 5 unidades.",
            "acao": "Faça uma entrada de estoque para estes produtos.",
            "produtos": [p.nome for p in relatorio["baixo_estoque"]]
        })
    
    # Verificar produtos com baixa margem
    if relatorio["baixa_margem"]:
        recomendacoes.append({
            "tipo": "alerta",
            "titulo": "📉 PRODUTOS COM BAIXA MARGEM",
            "mensagem": f"Você tem {len(relatorio['baixa_margem'])} produto(s) com margem < 20%.",
            "acao": "Considere aumentar o preço de venda ou revisar o custo.",
            "produtos": [p.nome for p in relatorio["baixa_margem"]]
        })

    if relatorio.get("produtos_abaixo_margem_alvo"):
        recomendacoes.append({
            "tipo": "alerta",
            "titulo": "🎯 PRODUTOS ABAIXO DA MARGEM ALVO",
            "mensagem": f"Você tem {len(relatorio['produtos_abaixo_margem_alvo'])} produto(s) abaixo da margem alvo.",
            "acao": "Ajuste preço de venda ou revise custo para atingir a margem definida.",
            "produtos": [p["nome"] for p in relatorio["produtos_abaixo_margem_alvo"][:10]]
        })
    
    # Sucesso: muitos produtos com boa margem
    if relatorio["alta_margem"]:
        recomendacoes.append({
            "tipo": "sucesso",
            "titulo": "✅ PRODUTOS COM BOA MARGEM",
            "mensagem": f"Parabéns! Você tem {len(relatorio['alta_margem'])} produto(s) com margem ≥ 50%.",
            "acao": "Continue com a estratégia de preço para estes produtos.",
            "produtos": [p.nome for p in relatorio["alta_margem"]]
        })
    
    # Info: crescimento
    if relatorio["total_clientes"] == 0:
        recomendacoes.append({
            "tipo": "info",
            "titulo": "📋 COMECE A CADASTRAR",
            "mensagem": "Nenhum cliente cadastrado ainda.",
            "acao": "Cadastre seus clientes na aba de Clientes.",
            "produtos": []
        })
    
    if relatorio["total_produtos"] == 0:
        recomendacoes.append({
            "tipo": "info",
            "titulo": "📦 COMECE COM PRODUTOS",
            "mensagem": "Nenhum produto cadastrado ainda.",
            "acao": "Cadastre seus produtos na aba de Estoque.",
            "produtos": []
        })
    
    # Dica: margem média
    if relatorio["total_produtos"] > 0:
        margem_media = relatorio["margem_media"]
        if margem_media < 30:
            recomendacoes.append({
                "tipo": "aviso",
                "titulo": "💡 MARGEM MÉDIA BAIXA",
                "mensagem": f"Sua margem média é de {margem_media:.1f}%, abaixo do ideal (30%+).",
                "acao": "Revise seus preços de venda ou custo de aquisição.",
                "produtos": []
            })
    
    return recomendacoes


def gerar_ranking_estoque(relatorio: dict) -> list:
    """
    Gera ranking de produtos por quantidade em estoque.
    
    Args:
        relatorio (dict): Relatório geral
    
    Returns:
        list: Produtos ordenados por quantidade
    """
    return sorted(relatorio["produtos"], key=lambda p: p.quantidade, reverse=True)


def gerar_ranking_margem(relatorio: dict) -> list:
    """
    Gera ranking de produtos por margem de lucro.
    
    Args:
        relatorio (dict): Relatório geral
    
    Returns:
        list: Produtos ordenados por margem
    """
    return sorted(relatorio["produtos"], key=lambda p: p.margem_lucro, reverse=True)


def exportar_relatorio_pdf(relatorio: dict, caminho: str = None) -> str:
    """
    Exporta o relatório para PDF.
    
    Args:
        relatorio (dict): Dados do relatório
        caminho (str): Caminho para salvar o arquivo (opcional)
    
    Returns:
        str: Caminho do arquivo gerado
    """
    if not caminho:
        # Criar diretório de relatórios se não existir
        relatorios_dir = Path("relatorios").absolute()
        relatorios_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = str(relatorios_dir / f"relatorio_{timestamp}.pdf")
    
    # Criar PDF
    doc = SimpleDocTemplate(caminho, pagesize=A4)
    elementos = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#0D47A1'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    # Título
    elementos.append(Paragraph("CASA GUARANI - RELATÓRIO GERENCIAL", title_style))
    elementos.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elementos.append(Spacer(1, 20))
    
    # Resumo Geral
    elementos.append(Paragraph("RESUMO GERAL", heading_style))
    
    resumo_data = [
        ['Métrica', 'Valor'],
        ['Total de Clientes', str(relatorio['total_clientes'])],
        ['Total de Produtos', str(relatorio['total_produtos'])],
        ['Itens em Estoque', str(relatorio['total_itens_estoque'])],
        ['Valor Total de Custo', f"R$ {relatorio['valor_total_custo']:,.2f}"],
        ['Valor Potencial de Venda', f"R$ {relatorio['valor_total_venda']:,.2f}"],
        ['Lucro Potencial', f"R$ {relatorio['lucro_potencial']:,.2f}"],
        ['Margem Média', f"{relatorio['margem_media']:.1f}%"],
    ]
    
    caixa = relatorio.get("caixa", {})
    caixa_atual = caixa.get("atual") or {}
    if caixa.get("aberto"):
        resumo_data.extend([
            ['Caixa Atual', 'Aberto'],
            ['Saldo Esperado do Caixa', f"R$ {caixa_atual.get('saldo_esperado', 0):,.2f}"],
            ['Entradas do Caixa', f"R$ {caixa_atual.get('total_entradas', 0):,.2f}"],
            ['Saidas do Caixa', f"R$ {caixa_atual.get('total_saidas', 0):,.2f}"],
        ])
    elif caixa.get("ultimo_fechado"):
        ultimo = caixa["ultimo_fechado"]
        resumo_data.extend([
            ['Caixa Atual', 'Fechado'],
            ['Ultimo Caixa Esperado', f"R$ {ultimo.get('saldo_esperado', 0):,.2f}"],
            ['Ultimo Caixa Contado', f"R$ {ultimo.get('saldo_contado', 0):,.2f}"],
            ['Diferenca do Ultimo Caixa', f"R$ {ultimo.get('diferenca', 0):,.2f}"],
        ])

    resumo_table = Table(resumo_data, colWidths=[3*inch, 2*inch])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elementos.append(resumo_table)
    elementos.append(Spacer(1, 20))
    
    if relatorio.get("caixa", {}).get("historico"):
        elementos.append(Paragraph("CAIXA", heading_style))
        caixa_data = [['ID', 'Abertura', 'Fechamento', 'Esperado', 'Contado', 'Diferenca']]
        for caixa_item in relatorio["caixa"]["historico"][:8]:
            caixa_data.append([
                str(caixa_item.get("id", "")),
                str(caixa_item.get("data_abertura") or "")[:16],
                str(caixa_item.get("data_fechamento") or "-")[:16],
                f"R$ {float(caixa_item.get('saldo_esperado') or caixa_item.get('saldo_inicial') or 0):,.2f}",
                f"R$ {float(caixa_item.get('saldo_contado') or 0):,.2f}" if caixa_item.get("saldo_contado") is not None else "-",
                f"R$ {float(caixa_item.get('diferenca') or 0):,.2f}" if caixa_item.get("diferenca") is not None else "-",
            ])
        caixa_table = Table(caixa_data, colWidths=[0.45*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
        caixa_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elementos.append(caixa_table)
        elementos.append(Spacer(1, 20))

    # Produtos em Estoque
    elementos.append(Paragraph("PRODUTOS EM ESTOQUE", heading_style))
    
    produtos_data = [['SKU', 'Nome', 'Qtd', 'Preço Venda', 'Margem %']]
    for produto in relatorio['produtos'][:20]:  # Limitar a 20 produtos
        produtos_data.append([
            produto.sku or '',
            produto.nome[:30],
            str(produto.quantidade),
            f"R$ {produto.preco_venda:.2f}",
            f"{produto.margem_lucro:.1f}%"
        ])
    
    produtos_table = Table(produtos_data, colWidths=[1*inch, 2.5*inch, 0.7*inch, 1*inch, 0.8*inch])
    produtos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFC107')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elementos.append(produtos_table)
    elementos.append(Spacer(1, 20))
    
    # Alertas
    if relatorio['baixo_estoque']:
        elementos.append(Paragraph("ALERTAS - ESTOQUE BAIXO", heading_style))
        alertas_data = [['Produto', 'Quantidade']]
        for produto in relatorio['baixo_estoque'][:10]:
            alertas_data.append([produto.nome[:40], str(produto.quantidade)])
        
        alertas_table = Table(alertas_data, colWidths=[4*inch, 1*inch])
        alertas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f44336')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elementos.append(alertas_table)

    # Produtos abaixo da margem alvo
    abaixo_alvo = relatorio.get("produtos_abaixo_margem_alvo", [])
    if abaixo_alvo:
        elementos.append(Spacer(1, 16))
        elementos.append(Paragraph("ALERTAS - MARGEM ABAIXO DO ALVO", heading_style))
        alvo_data = [['Produto', 'Categoria', 'Margem Atual', 'Margem Alvo']]
        for item in abaixo_alvo[:10]:
            alvo_data.append([
                item.get("nome", "")[:40],
                item.get("categoria", "Geral"),
                f"{item.get('margem_atual', 0):.1f}%",
                f"{item.get('margem_alvo', 0):.1f}%"
            ])

        alvo_table = Table(alvo_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1*inch])
        alvo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9800')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elementos.append(alvo_table)
    
    # Gerar PDF
    doc.build(elementos)
    return caminho


def exportar_relatorio_excel(relatorio: dict, caminho: str = None) -> str:
    """
    Exporta o relatório para Excel.
    
    Args:
        relatorio (dict): Dados do relatório
        caminho (str): Caminho para salvar o arquivo (opcional)
    
    Returns:
        str: Caminho do arquivo gerado
    """
    if not caminho:
        relatorios_dir = Path("relatorios").absolute()
        relatorios_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = str(relatorios_dir / f"relatorio_{timestamp}.xlsx")
    
    # Criar workbook
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="1976D2")
    alert_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ABA 1: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    ws_resumo['A1'] = "CASA GUARANI - RELATÓRIO GERENCIAL"
    ws_resumo['A1'].font = title_font
    ws_resumo.merge_cells('A1:B1')
    
    ws_resumo['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resumo.merge_cells('A2:B2')
    
    ws_resumo['A4'] = "Métrica"
    ws_resumo['B4'] = "Valor"
    ws_resumo['A4'].fill = header_fill
    ws_resumo['B4'].fill = header_fill
    ws_resumo['A4'].font = header_font
    ws_resumo['B4'].font = header_font
    
    resumo_dados = [
        ["Total de Clientes", relatorio['total_clientes']],
        ["Total de Produtos", relatorio['total_produtos']],
        ["Itens em Estoque", relatorio['total_itens_estoque']],
        ["Valor Total de Custo", f"R$ {relatorio['valor_total_custo']:,.2f}"],
        ["Valor Potencial de Venda", f"R$ {relatorio['valor_total_venda']:,.2f}"],
        ["Lucro Potencial", f"R$ {relatorio['lucro_potencial']:,.2f}"],
        ["Margem Média", f"{relatorio['margem_media']:.1f}%"],
    ]
    
    caixa = relatorio.get("caixa", {})
    caixa_atual = caixa.get("atual") or {}
    if caixa.get("aberto"):
        resumo_dados.extend([
            ["Caixa Atual", "Aberto"],
            ["Saldo Esperado do Caixa", f"R$ {caixa_atual.get('saldo_esperado', 0):,.2f}"],
            ["Entradas do Caixa", f"R$ {caixa_atual.get('total_entradas', 0):,.2f}"],
            ["Saidas do Caixa", f"R$ {caixa_atual.get('total_saidas', 0):,.2f}"],
        ])
    elif caixa.get("ultimo_fechado"):
        ultimo = caixa["ultimo_fechado"]
        resumo_dados.extend([
            ["Caixa Atual", "Fechado"],
            ["Ultimo Caixa Esperado", f"R$ {ultimo.get('saldo_esperado', 0):,.2f}"],
            ["Ultimo Caixa Contado", f"R$ {ultimo.get('saldo_contado', 0):,.2f}"],
            ["Diferenca do Ultimo Caixa", f"R$ {ultimo.get('diferenca', 0):,.2f}"],
        ])

    for i, (metrica, valor) in enumerate(resumo_dados, start=5):
        ws_resumo[f'A{i}'] = metrica
        ws_resumo[f'B{i}'] = valor
        ws_resumo[f'A{i}'].border = border
        ws_resumo[f'B{i}'].border = border
    
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 20

    ws_caixa = wb.create_sheet("Caixa")
    ws_caixa['A1'] = "FECHAMENTO DE CAIXA"
    ws_caixa['A1'].font = title_font
    ws_caixa.merge_cells('A1:G1')
    caixa_headers = ["ID", "Abertura", "Fechamento", "Esperado", "Contado", "Diferenca", "Status"]
    for col, header in enumerate(caixa_headers, start=1):
        cell = ws_caixa.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    for row, caixa_item in enumerate(relatorio.get("caixa", {}).get("historico", []), start=4):
        ws_caixa.cell(row=row, column=1).value = caixa_item.get("id")
        ws_caixa.cell(row=row, column=2).value = str(caixa_item.get("data_abertura") or "")[:16]
        ws_caixa.cell(row=row, column=3).value = str(caixa_item.get("data_fechamento") or "-")[:16]
        ws_caixa.cell(row=row, column=4).value = caixa_item.get("saldo_esperado") if caixa_item.get("saldo_esperado") is not None else caixa_item.get("saldo_inicial")
        ws_caixa.cell(row=row, column=5).value = caixa_item.get("saldo_contado")
        ws_caixa.cell(row=row, column=6).value = caixa_item.get("diferenca")
        ws_caixa.cell(row=row, column=7).value = caixa_item.get("status")
        for col in range(1, 8):
            ws_caixa.cell(row=row, column=col).border = border
    for col in "ABCDEFG":
        ws_caixa.column_dimensions[col].width = 18
    
    # ABA 2: Produtos
    ws_produtos = wb.create_sheet("Produtos")
    
    headers = ['SKU', 'Nome', 'Categoria', 'Quantidade', 'Preço Compra', 'Preço Venda', 'Margem %', 'Valor Estoque']
    for col, header in enumerate(headers, start=1):
        cell = ws_produtos.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row, produto in enumerate(relatorio['produtos'], start=2):
        ws_produtos.cell(row=row, column=1).value = produto.sku or ''
        ws_produtos.cell(row=row, column=2).value = produto.nome
        ws_produtos.cell(row=row, column=3).value = produto.categoria or ''
        ws_produtos.cell(row=row, column=4).value = produto.quantidade
        ws_produtos.cell(row=row, column=5).value = produto.preco_compra
        ws_produtos.cell(row=row, column=6).value = produto.preco_venda
        ws_produtos.cell(row=row, column=7).value = f"{produto.margem_lucro:.1f}%"
        ws_produtos.cell(row=row, column=8).value = produto.preco_venda * produto.quantidade
        
        # Aplicar borda
        for col in range(1, 9):
            ws_produtos.cell(row=row, column=col).border = border
        
        # Destacar baixo estoque
        if produto.quantidade <= 5:
            for col in range(1, 9):
                ws_produtos.cell(row=row, column=col).fill = alert_fill
    
    # Ajustar larguras
    ws_produtos.column_dimensions['A'].width = 12
    ws_produtos.column_dimensions['B'].width = 30
    ws_produtos.column_dimensions['C'].width = 15
    ws_produtos.column_dimensions['D'].width = 10
    ws_produtos.column_dimensions['E'].width = 12
    ws_produtos.column_dimensions['F'].width = 12
    ws_produtos.column_dimensions['G'].width = 10
    ws_produtos.column_dimensions['H'].width = 15
    
    # ABA 3: Alertas
    ws_alertas = wb.create_sheet("Alertas")
    
    ws_alertas['A1'] = "PRODUTOS COM ESTOQUE BAIXO"
    ws_alertas['A1'].font = title_font
    ws_alertas.merge_cells('A1:C1')
    
    ws_alertas['A3'] = "Produto"
    ws_alertas['B3'] = "Quantidade"
    ws_alertas['C3'] = "Preço Venda"
    
    for col in range(1, 4):
        cell = ws_alertas.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
    
    for row, produto in enumerate(relatorio['baixo_estoque'], start=4):
        ws_alertas.cell(row=row, column=1).value = produto.nome
        ws_alertas.cell(row=row, column=2).value = produto.quantidade
        ws_alertas.cell(row=row, column=3).value = f"R$ {produto.preco_venda:.2f}"
        
        for col in range(1, 4):
            ws_alertas.cell(row=row, column=col).border = border
            ws_alertas.cell(row=row, column=col).fill = alert_fill
    
    ws_alertas.column_dimensions['A'].width = 40
    ws_alertas.column_dimensions['B'].width = 12
    ws_alertas.column_dimensions['C'].width = 15

    # ABA 4: Margem Alvo
    abaixo_alvo = relatorio.get("produtos_abaixo_margem_alvo", [])
    ws_margem = wb.create_sheet("Margem Alvo")

    ws_margem['A1'] = "PRODUTOS ABAIXO DA MARGEM ALVO"
    ws_margem['A1'].font = title_font
    ws_margem.merge_cells('A1:D1')

    ws_margem['A3'] = "Produto"
    ws_margem['B3'] = "Categoria"
    ws_margem['C3'] = "Margem Atual"
    ws_margem['D3'] = "Margem Alvo"

    for col in range(1, 5):
        cell = ws_margem.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for row, item in enumerate(abaixo_alvo, start=4):
        ws_margem.cell(row=row, column=1).value = item.get("nome", "")
        ws_margem.cell(row=row, column=2).value = item.get("categoria", "Geral")
        ws_margem.cell(row=row, column=3).value = f"{item.get('margem_atual', 0):.1f}%"
        ws_margem.cell(row=row, column=4).value = f"{item.get('margem_alvo', 0):.1f}%"

        for col in range(1, 5):
            ws_margem.cell(row=row, column=col).border = border
            ws_margem.cell(row=row, column=col).fill = alert_fill

    ws_margem.column_dimensions['A'].width = 40
    ws_margem.column_dimensions['B'].width = 20
    ws_margem.column_dimensions['C'].width = 15
    ws_margem.column_dimensions['D'].width = 15
    
    # Salvar
    wb.save(caminho)
    return caminho

